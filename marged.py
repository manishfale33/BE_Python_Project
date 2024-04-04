import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import simpledialog
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
import csv
import random

# Function to send a message using WhatsApp
def send_whatsapp_message():
    selected_item = whatsapp_phone_tree.selection()
    message = whatsapp_message_entry.get()
    if not message:
        message = generate_random_message()

    if selected_item:
        item = whatsapp_phone_tree.item(selected_item)
        phone_number = item['values'][1]
        send_message(driver, phone_number, message)

# Function to send a message using Facebook
def send_facebook_message_with_attachment():
    username = facebook_username_entry.get()
    password = facebook_password_entry.get()
    recipient = facebook_recipient_entry.get()
    message_text = facebook_message_entry.get("1.0", tk.END)
    attachment_path = facebook_attachment_entry.get()

    driver = webdriver.Chrome()

    try:
        driver.get("https://www.facebook.com")

        email_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "email")))
        email_field.send_keys(username)

        password_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "pass")))
        password_field.send_keys(password)

        login_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "login")))
        login_button.click()

        WebDriverWait(driver, 50).until(EC.url_contains("https://www.facebook.com/"))

        if "login" in driver.current_url:
            messagebox.showerror("Error", "Login failed. Please check your credentials.")
            return

        search_box = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='search']")))
        search_box.send_keys(recipient)
        search_box.send_keys(Keys.RETURN)

        handle_notifications(driver)

        message_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//a[contains(@aria-label, 'Message')]")))
        message_button.click()

        message_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[contenteditable='true']")))
        message_input.send_keys(message_text)

        if attachment_path:
            attachment_input = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
            attachment_input.send_keys(attachment_path)

        message_input.send_keys(Keys.RETURN)

        messagebox.showinfo("Success", "Message sent successfully")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

    finally:
        driver.quit()

# Function to send an email using Gmail
def send_gmail_email():
    sender_email = gmail_sender_email_entry.get()
    sender_password = gmail_sender_password_entry.get()
    recipients = gmail_recipients_text.get("1.0", "end-1c").split('\n')
    subject = gmail_subject_entry.get()
    message_text = gmail_message_entry.get("1.0", "end-1c")

    driver = webdriver.Chrome()

    try:
        driver.get("https://mail.google.com")

        email_field = driver.find_element(By.ID, "identifierId")
        email_field.send_keys(sender_email)
        email_field.send_keys(Keys.RETURN)

        driver.implicitly_wait(10)
        password_field = driver.find_element(By.NAME, "password")
        password_field.send_keys(sender_password)
        password_field.send_keys(Keys.RETURN)

        compose_button = driver.find_element(By.CSS_SELECTOR, "div[role='button'][gh='cm']")
        compose_button.click()
        driver.implicitly_wait(5)

        to_field = driver.find_element_by_name("to")
        to_field.send_keys(", ".join(recipients))
        to_field.send_keys(Keys.RETURN)

        subject_field = driver.find_element(By.NAME, "subjectbox")
        subject_field.send_keys(subject)

        message_field = driver.find_element(By.CSS_SELECTOR, "div[role='textbox']")
        message_field.send_keys(message_text)

        send_button = driver.find_element(By.CSS_SELECTOR, "div[aria-label='Send ‪(Ctrl-Enter)‬']")
        send_button.click()

        messagebox.showinfo("Success", "Email sent successfully")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

    finally:
        driver.quit()

# Function to send a message using Instagram
def send_instagram_message():
    # Placeholder for sending messages via Instagram API
    messagebox.showinfo("Instagram", "Functionality not implemented yet.")

# Function to fetch phone numbers and contact names from CSV
def fetch_phone_contacts_from_csv(file_path):
    phone_contacts = []
    with open(file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            if row:
                if len(row) == 2:
                    phone_contacts.append([row[0], row[1]])
                elif len(row) == 1:
                    phone_contacts.append(["", row[0]])
    return phone_contacts

# Function to fetch phone numbers and contact names from Excel
def fetch_phone_contacts_from_excel(file_path):
    phone_contacts = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        if row:
            if len(row) == 2:
                phone_contacts.append([row[0], row[1]])
            elif len(row) == 1:
                phone_contacts.append(["", str(row[0])])
    return phone_contacts

# Function to generate a random message
def generate_random_message():
    messages = ["Hello!", "How are you?", "Good morning!", "Happy to connect!"]
    return random.choice(messages)

# Function to handle file selection and phone number fetching for WhatsApp
def select_whatsapp_file_and_fetch_numbers():
    file_path = filedialog.askopenfilename()
    whatsapp_phone_tree.delete(*whatsapp_phone_tree.get_children())
    
    if file_path.endswith('.csv'):
        phone_contacts = fetch_phone_contacts_from_csv(file_path)
    elif file_path.endswith(('.xlsx', '.xls')):
        phone_contacts = fetch_phone_contacts_from_excel(file_path)
    
    for contact in phone_contacts:
        whatsapp_phone_tree.insert('', 'end', values=contact)

# Function to handle file selection and phone number fetching for Facebook
def select_facebook_file_and_fetch_numbers():
    file_path = filedialog.askopenfilename()
    facebook_phone_tree.delete(*facebook_phone_tree.get_children())
    
    if file_path.endswith('.csv'):
        phone_contacts = fetch_phone_contacts_from_csv(file_path)
    elif file_path.endswith(('.xlsx', '.xls')):
        phone_contacts = fetch_phone_contacts_from_excel(file_path)
    
    for contact in phone_contacts:
        facebook_phone_tree.insert('', 'end', values=contact)

# Function to handle file selection and email fetching for Gmail
def select_gmail_file_and_fetch_emails():
    file_path = filedialog.askopenfilename()
    gmail_recipients_text.delete(1.0, tk.END)
    
    if file_path.endswith('.csv'):
        with open(file_path, 'r') as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                if row:
                    gmail_recipients_text.insert(tk.END, row[0] + '\n')
    elif file_path.endswith(('.xlsx', '.xls')):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row:
                gmail_recipients_text.insert(tk.END, row[0] + '\n')

# Create the main window
root = tk.Tk()
root.title("Multi-Messaging App")

# Create tab control
tab_control = ttk.Notebook(root)

# WhatsApp Tab
whatsapp_tab = ttk.Frame(tab_control)
tab_control.add(whatsapp_tab, text="WhatsApp")
tab_control.pack(expand=1, fill="both")

whatsapp_message_label = ttk.Label(whatsapp_tab, text="Message (leave blank for random):")
whatsapp_message_label.pack()
whatsapp_message_entry = ttk.Entry(whatsapp_tab)
whatsapp_message_entry.pack()

whatsapp_select_file_button = ttk.Button(whatsapp_tab, text="Select CSV/Excel File", command=select_whatsapp_file_and_fetch_numbers)
whatsapp_select_file_button.pack()

whatsapp_phone_tree = ttk.Treeview(whatsapp_tab, columns=("Name", "Phone"))
whatsapp_phone_tree.heading("#1", text="Name")
whatsapp_phone_tree.heading("#2", text="Phone")
whatsapp_phone_tree.pack()

whatsapp_send_button = ttk.Button(whatsapp_tab, text="Send Message", command=send_whatsapp_message)
whatsapp_send_button.pack()

# Facebook Tab
facebook_tab = ttk.Frame(tab_control)
tab_control.add(facebook_tab, text="Facebook")

facebook_username_label = ttk.Label(facebook_tab, text="Facebook Email/Phone:")
facebook_username_label.pack()
facebook_username_entry = ttk.Entry(facebook_tab)
facebook_username_entry.pack()

facebook_password_label = ttk.Label(facebook_tab, text="Facebook Password:")
facebook_password_label.pack()
facebook_password_entry = ttk.Entry(facebook_tab, show='*')
facebook_password_entry.pack()

facebook_recipient_label = ttk.Label(facebook_tab, text="Recipient's Facebook Name:")
facebook_recipient_label.pack()
facebook_recipient_entry = ttk.Entry(facebook_tab)
facebook_recipient_entry.pack()

facebook_message_label = ttk.Label(facebook_tab, text="Message:")
facebook_message_label.pack()
facebook_message_entry = tk.Text(facebook_tab, height=5, width=40)
facebook_message_entry.pack()

facebook_attachment_label = ttk.Label(facebook_tab, text="Attachment Path (optional):")
facebook_attachment_label.pack()
facebook_attachment_entry = ttk.Entry(facebook_tab)
facebook_attachment_entry.pack()

facebook_select_file_button = ttk.Button(facebook_tab, text="Select CSV/Excel File", command=select_facebook_file_and_fetch_numbers)
facebook_select_file_button.pack()

facebook_phone_tree = ttk.Treeview(facebook_tab, columns=("Name", "Phone"))
facebook_phone_tree.heading("#1", text="Name")
facebook_phone_tree.heading("#2", text="Phone")
facebook_phone_tree.pack()

facebook_send_button = ttk.Button(facebook_tab, text="Send Message with Attachment", command=send_facebook_message_with_attachment)
facebook_send_button.pack()

# Gmail Tab
gmail_tab = ttk.Frame(tab_control)
tab_control.add(gmail_tab, text="Gmail")

gmail_sender_email_label = ttk.Label(gmail_tab, text="Your Email:")
gmail_sender_email_label.pack()
gmail_sender_email_entry = ttk.Entry(gmail_tab)
gmail_sender_email_entry.pack()

gmail_sender_password_label = ttk.Label(gmail_tab, text="Your Password:")
gmail_sender_password_label.pack()
gmail_sender_password_entry = ttk.Entry(gmail_tab, show='*')
gmail_sender_password_entry.pack()

gmail_recipients_label = ttk.Label(gmail_tab, text="Recipients (One per line):")
gmail_recipients_label.pack()
gmail_recipients_text = tk.Text(gmail_tab, height=10, width=40)
gmail_recipients_text.pack()

gmail_subject_label = ttk.Label(gmail_tab, text="Email Subject:")
gmail_subject_label.pack()
gmail_subject_entry = ttk.Entry(gmail_tab)
gmail_subject_entry.pack()

gmail_message_label = ttk.Label(gmail_tab, text="Email Message:")
gmail_message_label.pack()
gmail_message_entry = tk.Text(gmail_tab, height=5, width=40)
gmail_message_entry.pack()

gmail_select_file_button = ttk.Button(gmail_tab, text="Select CSV/Excel File", command=select_gmail_file_and_fetch_emails)
gmail_select_file_button.pack()

gmail_send_button = ttk.Button(gmail_tab, text="Send Email", command=send_gmail_email)
gmail_send_button.pack()

# Instagram Tab
instagram_tab = ttk.Frame(tab_control)
tab_control.add(instagram_tab, text="Instagram")

instagram_username_label = ttk.Label(instagram_tab, text="Instagram Username:")
instagram_username_label.pack()
instagram_username_entry = ttk.Entry(instagram_tab)
instagram_username_entry.pack()

instagram_password_label = ttk.Label(instagram_tab, text="Instagram Password:")
instagram_password_label.pack()
instagram_password_entry = ttk.Entry(instagram_tab, show='*')
instagram_password_entry.pack()

instagram_recipient_label = ttk.Label(instagram_tab, text="Recipient's Instagram Username:")
instagram_recipient_label.pack()
instagram_recipient_entry = ttk.Entry(instagram_tab)
instagram_recipient_entry.pack()

instagram_message_label = ttk.Label(instagram_tab, text="Message:")
instagram_message_label.pack()
instagram_message_entry = tk.Text(instagram_tab, height=5, width=40)
instagram_message_entry.pack()

instagram_send_button = ttk.Button(instagram_tab, text="Send Message", command=send_instagram_message)
instagram_send_button.pack()

# Start the GUI event loop
root.mainloop()
