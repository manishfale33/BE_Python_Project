import time
import csv
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk, filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options

class FacebookMessageSender:
    def __init__(self, frame):
        self.frame = frame

        # Define custom colors
        dark_background = "#000000"
        text_color = "#00A0E6"

        # Create labels and entry fields for username, password, message, and attachment
        self.username_label = Label(frame, text="Facebook Email/Phone:", fg=text_color, bg=dark_background)
        self.username_label.grid(row=0, column=0, sticky=W)
        self.username_entry = ttk.Entry(frame, style="Custom.TEntry")  # Use themed entry widget
        self.username_entry.grid(row=0, column=1, sticky=W)

        self.password_label = Label(frame, text="Facebook Password:", fg=text_color, bg=dark_background)
        self.password_label.grid(row=1, column=0, sticky=W)
        self.password_entry = ttk.Entry(frame, show="*", style="Custom.TEntry")  # Use themed entry widget
        self.password_entry.grid(row=1, column=1, sticky=W)

        self.message_label = Label(frame, text="Message:", fg=text_color, bg=dark_background)
        self.message_label.grid(row=2, column=0, sticky=W)
        self.message_entry = Text(frame, height=5, width=40, bg=dark_background, fg=text_color)
        self.message_entry.grid(row=2, column=1, columnspan=2, sticky=W)

        self.attachment_label = Label(frame, text="Attachment Path (optional):", fg=text_color, bg=dark_background)
        self.attachment_label.grid(row=3, column=0, sticky=W)
        self.attachment_entry = ttk.Entry(frame, style="Custom.TEntry")  # Use themed entry widget
        self.attachment_entry.grid(row=3, column=1, sticky=W)

        self.attach_label = Label(frame, text="Attachment:", fg=text_color, bg=dark_background)
        self.attach_label.grid(row=4, column=0, sticky=W)

        # Create browse button for attachment
        self.attachment_button = PhotoImage(file="Images/search.png").subsample(10)  # Resized image
        self.browse_button = Button(frame, image=self.attachment_button, command=self.browse_for_attachment, borderwidth=0, bg=dark_background)
        self.browse_button.grid(row=4, column=1, sticky=E)

        # Create send button for sending message or friend request
        self.send_image = PhotoImage(file="Images/send.png").subsample(10)  # Resized image
        self.send_button = Button(frame, image=self.send_image, command=self.send_message, borderwidth=0, bg=dark_background)
        self.send_button.grid(row=4, column=2, sticky=W)

        # Add a button to select recipients from a file
        self.select_recipients_button = Button(frame, text="Select Recipients from File", command=self.select_recipients_from_file, bg=dark_background, fg=text_color, relief=FLAT)
        self.select_recipients_button.grid(row=5, column=0, columnspan=2, sticky=W)

        # Add a text box to show the selected recipients
        self.selected_recipients_text = Text(frame, height=5, width=40, bg=dark_background, fg=text_color)
        self.selected_recipients_text.grid(row=5, column=1, columnspan=2, sticky=W)

        # Create a custom style for themed entry widgets
        self.custom_style = ttk.Style()
        self.custom_style.configure("Custom.TEntry", foreground=text_color, background=dark_background)

        # Make the application responsive
        for i in range(6):
            frame.rowconfigure(i, weight=1)
        for i in range(3):
            frame.columnconfigure(i, weight=1)

    def browse_for_attachment(self):
        # Function to browse and select a file for attachment
        file_path = filedialog.askopenfilename()
        self.attachment_entry.delete(0, END)
        self.attachment_entry.insert(0, file_path)

    def select_recipients_from_file(self):
        # Function to browse and select a CSV or Excel file containing the recipient list
        file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx")])
        recipient_list = self.read_recipient_list(file_path)
        if recipient_list:
            self.selected_recipients_text.delete("1.0", END)
            self.selected_recipients_text.insert(END, "\n".join(recipient_list))

    def read_recipient_list(self, file_path):
        # Function to read the recipient list from a CSV or Excel file
        recipient_list = []
        if file_path:
            if file_path.endswith(".csv"):
                with open(file_path, 'r') as file:
                    reader = csv.reader(file)
                    for row in reader:
                        recipient_list.append(row[0])  # Assuming the first column contains the recipient names
            elif file_path.endswith(".xlsx"):
                wb = load_workbook(file_path)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    recipient_list.append(row[0])
        return recipient_list

    def send_message(self):
        # Function to send a message on Facebook
        username = self.username_entry.get()
        password = self.password_entry.get()
        message_text = self.message_entry.get("1.0", END).strip()
        attachment_path = self.attachment_entry.get()
        recipient_list = self.selected_recipients_text.get("1.0", END).strip().split("\n")

        print("Starting the message sending process...")

        # Chrome WebDriver options with --disable-notifications argument
        chrome_options = Options()
        chrome_options.add_argument("--disable-notifications")
        driver = webdriver.Chrome(options=chrome_options)

        try:
            print("Opening Facebook login page...")
            driver.get("https://www.facebook.com")

            print("Entering username...")
            email_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "email")))
            email_field.send_keys(username)

            print("Entering password...")
            password_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "pass")))
            password_field.send_keys(password)

            print("Logging in...")
            login_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "login")))
            login_button.click()

            WebDriverWait(driver, 50).until(EC.url_contains("https://www.facebook.com/"))

            if "login" in driver.current_url:
                print("Login failed. Please check your credentials.")
                return

            print("Logged in successfully.")

            for recipient in recipient_list:
                # Search for the recipient by name
                print(f"Searching for '{recipient}'...")
                search_box = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='search']")))
                search_box.send_keys(recipient)
                time.sleep(2)  # Adding a delay for search results to load
                search_box.send_keys(Keys.RETURN)

                # Handle site notifications (if they appear)
                self.handle_notifications(driver)

                # Click on the profile icon of the searched person
                profile_icon_xpath = "//body[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/h2[1]/span[1]/span[1]/span[1]/a[1]/span[1]"
                profile_icon = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, profile_icon_xpath)))
                profile_icon.click()

                # Click the message button
                message_button_xpath = "//span[contains(text(),'Message')]"
                message_button = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, message_button_xpath)))
                message_button.click()

                # Wait for the message box to load
                message_box = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body._6s5d._71pn.system-fonts--body.segoe:nth-child(2) div.x9f619.x1n2onr6.x1ja2u2z div.x1ey2m1c.xds687c.xixxii4:nth-child(1) div.xuk3077.x78zum5.xc8icb0:nth-child(1) div.x1ey2m1c.x78zum5.x164qtfw.xixxii4.x1vjfegm:nth-child(1) div.x9f619.x1n2onr6.x1ja2u2z.__fb-light-mode.x78zum5.xdt5ytf.x1iyjqo2.xs83m0k.x193iq5w div.x1uvtmcs.x4k7w5x.x1h91t0o.x1beo9mf.xaigb6o.x12ejxvf.x3igimt.xarpa2k.xedcshv.x1lytzrv.x1t2pt76.x7ja8zs.x1n2onr6.x1qrby5j.x1jfb8zj div.x5yr21d.x1uvtmcs div.xcrg951.xgqcy7u.x1lq5wgf.x78zum5.x6prxxf.xvq8zen.x17adc0v.xi55695.x1rgmuzj.xbbk1sx.x6l8u58 div.x78zum5.xdt5ytf.x1iyjqo2.x193iq5w.x2lwn1j.x1n2onr6:nth-child(2) div.xuk3077.x57kliw.x78zum5.x6prxxf.xz9dl7a.xsag5q8 div.x1iyjqo2.xw2csxc.x1n2onr6:nth-child(2) div.x78zum5.x1iyjqo2.x6q2ic0:nth-child(4) div.x16sw7j7.x107yiy2.xv8uw2v.x1tfwpuw.x2g32xy.x9f619.xlai7qp.x1iyjqo2.xeuugli div.x78zum5.x13a6bvl div.x78zum5.x1iyjqo2.xq8finb.x16n37ib.x1xmf6yo.x1e56ztr.xeuugli.x1n2onr6 div.xzsf02u.x1a2a7pz.x1n2onr6.x14wi4xw.x1iyjqo2.x1gh3ibb.xisnujt.xeuugli.x1odjw0f.notranslate > p.xat24cr.xdj266r")))

                # Enter the message directly into the message box
                print("Typing message...")
                message_box.send_keys(message_text)

                # Attach a file if specified
                if attachment_path:
                    attachment_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
                    attachment_input.send_keys(attachment_path)
                    time.sleep(2)  # Adding a delay for the attachment to upload

                # Click the send button
                print("Sending the message...")
                send_button = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//body[1]/div[1]/div[1]/div[1]/div[1]/div[5]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]/div[1]/span[2]/div[1]")))
                send_button.click()
                print("Message sent successfully")
                # Wait for some time before sending the next message
                time.sleep(5)  # Wait for 5 seconds

        except Exception as e:
            print(f"An error occurred: {str(e)}")

        finally:
            driver.quit()

    def handle_notifications(self, driver):
        # Function to handle Facebook site notifications
        try:
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'Not Now')]"))).click()
        except:
            pass

class WhatsAppMessageSender:
    def __init__(self, frame):
        self.frame = frame

        # Define custom colors
        dark_background = "#000000"
        text_color = "#00A0E6"

        # Create labels and entry fields for phone number and message
        self.phone_label = Label(frame, text="Phone Number:", fg=text_color, bg=dark_background)
        self.phone_label.grid(row=0, column=0, sticky=W)
        self.phone_entry = ttk.Entry(frame, style="Custom.TEntry")  # Use themed entry widget
        self.phone_entry.grid(row=0, column=1, sticky=W)

        self.message_label = Label(frame, text="Message:", fg=text_color, bg=dark_background)
        self.message_label.grid(row=1, column=0, sticky=W)
        self.message_entry = Text(frame, height=5, width=40, bg=dark_background, fg=text_color)
        self.message_entry.grid(row=1, column=1, columnspan=2, sticky=W)

        # Create send button for sending message
        self.send_image = PhotoImage(file="Images/send.png").subsample(10)  # Resized image
        self.send_button = Button(frame, image=self.send_image, command=self.send_message, borderwidth=0, bg=dark_background)
        self.send_button.grid(row=2, column=0, columnspan=2, sticky=W)

        # Create a custom style for themed entry widgets
        self.custom_style = ttk.Style()
        self.custom_style.configure("Custom.TEntry", foreground=text_color, background=dark_background)

        # Make the application responsive
        for i in range(3):
            frame.rowconfigure(i, weight=1)
        for i in range(3):
            frame.columnconfigure(i, weight=1)

    def send_message(self):
        # Function to send a message on WhatsApp
        phone_number = self.phone_entry.get()
        message_text = self.message_entry.get("1.0", END).strip()

        print("Starting the message sending process...")

        # Chrome WebDriver options with --disable-notifications argument
        chrome_options = Options()
        chrome_options.add_argument("--disable-notifications")
        driver = webdriver.Chrome(options=chrome_options)

        try:
            print("Opening WhatsApp Web...")
            driver.get("https://web.whatsapp.com/")

            print("Waiting for QR code scan...")
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'B8a')]/div/img")))

            print("QR code scanned. Sending message...")
            # Find the input field for phone number
            phone_input = driver.find_element(By.XPATH, "//div[contains(@class, 'copyable-text selectable-text')][@contenteditable='true']")
            phone_input.send_keys(phone_number)
            time.sleep(2)  # Adding a delay to let the contact load
            phone_input.send_keys(Keys.ENTER)

            # Find the input field for message
            message_input = driver.find_element(By.XPATH, "//div[contains(@class, 'copyable-text selectable-text')][@contenteditable='true']")
            message_input.send_keys(message_text)
            message_input.send_keys(Keys.ENTER)

            print("Message sent successfully")

        except Exception as e:
            print(f"An error occurred: {str(e)}")

        finally:
            driver.quit()

def main():
    root = Tk()
    tab_control = ttk.Notebook(root)
    facebook_tab = ttk.Frame(tab_control)
    whatsapp_tab = ttk.Frame(tab_control)
    tab_control.add(facebook_tab, text="Facebook")
    tab_control.add(whatsapp_tab, text="WhatsApp")
    tab_control.pack(expand=1, fill="both")
    FacebookMessageSender(facebook_tab)
    WhatsAppMessageSender(whatsapp_tab)
    root.mainloop()

if __name__ == "__main__":
    main()
